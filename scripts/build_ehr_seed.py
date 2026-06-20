import json
from openpyxl import Workbook
from openpyxl.styles import Font

COLUMNS = [
    "name", "slug", "logo", "website", "category",
    "rating", "reportUrl",
    "introduction", "ourVerdict",
    "keyTakeaways", "pros", "cons",
    "pictures",
    "howItWorks", "whoIsItFor", "howItIsDifferent", "sentiments",
    "specifications", "faqs",
]

def slugify(name):
    return name.lower().replace(" ", "-").replace("/", "-").replace("(", "").replace(")", "")

def arr(*items):
    return "|".join(items)

ROWS = [
    dict(
        name="Epic", website="https://www.epic.com", category="EHR/EMR",
        introduction="Epic is a comprehensive electronic health record (EHR) platform used primarily by large hospitals and health systems to manage patient records, scheduling, billing, and clinical workflows.",
        ourVerdict="Epic is the dominant EHR for large hospital systems thanks to its depth and interoperability, but its cost and implementation complexity put it out of reach for smaller practices.",
        keyTakeaways=arr("Widely used by large hospital systems and academic medical centers",
                          "Strong interoperability via Care Everywhere network",
                          "High implementation cost and long deployment timelines"),
        pros=arr("Deep clinical functionality across specialties", "Strong data exchange between Epic sites", "Robust patient portal (MyChart)"),
        cons=arr("Very high cost of ownership", "Complex, lengthy implementation", "Steep learning curve for staff"),
        whoIsItFor="Large hospitals, health systems, and academic medical centers needing enterprise-scale EHR.",
        howItIsDifferent="Focuses on enterprise-scale interoperability and a single unified database across the health system rather than modular point solutions.",
        specifications=json.dumps({"Deployment": "Cloud/On-Premise", "Best For": "Large Health Systems", "Pricing Model": "Custom/Enterprise"}),
    ),
    dict(
        name="Oracle Health (Cerner)", website="https://www.oracle.com/health/", category="EHR/EMR",
        introduction="Oracle Health, formerly Cerner, provides EHR and health IT solutions used by hospitals and health networks for clinical documentation, revenue cycle, and population health.",
        ourVerdict="A strong enterprise EHR backed by Oracle's cloud infrastructure, well suited to hospitals already invested in the Cerner ecosystem.",
        keyTakeaways=arr("Now integrated with Oracle Cloud Infrastructure", "Popular among mid-to-large hospitals", "Broad suite covering clinical and revenue cycle needs"),
        pros=arr("Strong hospital-grade clinical tools", "Backed by Oracle's cloud scale", "Good population health features"),
        cons=arr("Interface considered dated by some users", "Customization can require vendor support", "Significant training required"),
        whoIsItFor="Mid-to-large hospitals and health networks.",
        howItIsDifferent="Combines a legacy hospital EHR base with Oracle's cloud and data infrastructure investment.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Hospitals & Health Networks", "Pricing Model": "Custom/Enterprise"}),
    ),
    dict(
        name="athenahealth", website="https://www.athenahealth.com", category="EHR/EMR",
        introduction="athenahealth offers a cloud-based EHR and practice management platform aimed at ambulatory practices, combining clinical documentation, billing, and patient engagement tools.",
        ourVerdict="athenahealth is a solid choice for ambulatory practices wanting a cloud-native EHR with strong revenue cycle services bundled in.",
        keyTakeaways=arr("Cloud-native, no local servers required", "Includes built-in revenue cycle management services", "Popular with small and mid-size practices"),
        pros=arr("Easy to deploy and maintain", "Strong billing/RCM services", "Regular automatic updates"),
        cons=arr("Customization options are limited", "Reporting can be less flexible than competitors", "Support response times vary"),
        whoIsItFor="Ambulatory and outpatient practices of small to mid size.",
        howItIsDifferent="Bundles RCM/billing services directly with the EHR rather than treating them as a separate add-on.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Ambulatory Practices", "Pricing Model": "Subscription + % of collections"}),
    ),
    dict(
        name="eClinicalWorks", website="https://www.eclinicalworks.com", category="EHR/EMR",
        introduction="eClinicalWorks is a widely used EHR and practice management solution for ambulatory practices, offering clinical documentation, scheduling, billing, and telehealth.",
        ourVerdict="eClinicalWorks offers strong value for ambulatory practices needing a full-featured, configurable EHR at a competitive price point.",
        keyTakeaways=arr("Large install base among ambulatory practices", "Built-in telehealth and patient engagement tools", "Configurable templates by specialty"),
        pros=arr("Competitive pricing", "Extensive specialty-specific templates", "Integrated telehealth"),
        cons=arr("Interface can feel cluttered", "Past compliance/legal issues raised concerns", "Onboarding can take time"),
        whoIsItFor="Small to mid-size ambulatory practices across many specialties.",
        howItIsDifferent="Offers a large library of specialty-specific templates at a lower price point than enterprise EHRs.",
        specifications=json.dumps({"Deployment": "Cloud/On-Premise", "Best For": "Ambulatory Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="NextGen Healthcare", website="https://www.nextgen.com", category="EHR/EMR",
        introduction="NextGen Healthcare provides EHR and practice management software for ambulatory practices, with tools for clinical documentation, billing, and population health management.",
        ourVerdict="NextGen is a flexible EHR option for ambulatory and specialty practices that want strong interoperability and population health tools.",
        keyTakeaways=arr("Strong focus on ambulatory and specialty care", "Population health and analytics tools included", "Supports value-based care workflows"),
        pros=arr("Good interoperability features", "Specialty-specific workflows", "Population health analytics"),
        cons=arr("Pricing can be high for small practices", "Some users report slower customer support", "UI updates needed in older modules"),
        whoIsItFor="Ambulatory and specialty practices focused on value-based care.",
        howItIsDifferent="Places heavier emphasis on population health and value-based care analytics than many competitors.",
        specifications=json.dumps({"Deployment": "Cloud/On-Premise", "Best For": "Specialty Ambulatory Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="Veradigm (Allscripts)", website="https://veradigm.com", category="EHR/EMR",
        introduction="Veradigm, formerly Allscripts, offers EHR solutions for ambulatory and hospital settings along with data and analytics services for healthcare organizations.",
        ourVerdict="Veradigm remains a capable EHR for practices that also want access to its broader health data and analytics network.",
        keyTakeaways=arr("Rebranded from Allscripts to Veradigm", "Offers both EHR and health data/analytics products", "Used across ambulatory and hospital settings"),
        pros=arr("Broad portfolio including data analytics", "Established presence in healthcare IT", "Multiple product tiers for different practice sizes"),
        cons=arr("Brand transition has caused some user confusion", "Interface considered less modern by some reviewers", "Support quality varies by product line"),
        whoIsItFor="Ambulatory practices and hospitals wanting EHR plus data/analytics services.",
        howItIsDifferent="Pairs its EHR with a separate health data and analytics business, unlike pure-EHR competitors.",
        specifications=json.dumps({"Deployment": "Cloud/On-Premise", "Best For": "Ambulatory & Hospital", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="MEDITECH Expanse", website="https://ehr.meditech.com", category="EHR/EMR",
        introduction="MEDITECH Expanse is a cloud-based EHR platform designed for hospitals and health systems, offering clinical, financial, and patient engagement modules in one system.",
        ourVerdict="MEDITECH Expanse is a strong fit for community hospitals seeking an integrated, cost-effective alternative to Epic or Oracle Health.",
        keyTakeaways=arr("Popular among community and rural hospitals", "Single-database architecture across modules", "Generally lower cost than Epic/Oracle Health"),
        pros=arr("More affordable than top-tier enterprise EHRs", "Unified single-database design", "Strong fit for community hospitals"),
        cons=arr("Smaller third-party app ecosystem", "Less brand recognition than Epic/Cerner", "Customization requests can be slower"),
        whoIsItFor="Community hospitals and smaller health systems.",
        howItIsDifferent="Targets the community hospital market with a lower-cost, single-database alternative to enterprise giants.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Community Hospitals", "Pricing Model": "Custom/Enterprise"}),
    ),
    dict(
        name="Practice Fusion", website="https://www.practicefusion.com", category="EHR/EMR",
        introduction="Practice Fusion is a cloud-based EHR aimed at small independent practices, offering charting, e-prescribing, and scheduling at a low entry cost.",
        ourVerdict="Practice Fusion suits small independent practices wanting an affordable, easy-to-use EHR without heavy customization needs.",
        keyTakeaways=arr("Targets small independent practices", "Low-cost cloud-based EHR", "Owned by Veradigm"),
        pros=arr("Affordable pricing for small practices", "Easy to set up and use", "Built-in e-prescribing"),
        cons=arr("Limited advanced/specialty features", "Fewer integrations than larger platforms", "Support tiers vary by plan"),
        whoIsItFor="Small independent practices and solo practitioners.",
        howItIsDifferent="Prioritizes simplicity and low cost over the deep specialty customization of larger EHRs.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Small Independent Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="DrChrono", website="https://www.drchrono.com", category="EHR/EMR",
        introduction="DrChrono is a cloud and mobile-friendly EHR and practice management platform built for small to mid-size medical practices.",
        ourVerdict="DrChrono is a good fit for mobile-first, small to mid-size practices that value an iPad-friendly charting experience.",
        keyTakeaways=arr("Strong mobile/iPad-based charting experience", "Targets small to mid-size practices", "Includes billing and patient engagement tools"),
        pros=arr("Mobile-friendly, iPad-native app", "Customizable forms and templates", "Integrated billing"),
        cons=arr("Some advanced features require higher-tier plans", "Occasional sync issues reported by users", "Less suited to very large organizations"),
        whoIsItFor="Small to mid-size practices wanting a mobile-first EHR.",
        howItIsDifferent="Built mobile-first around iPad/tablet charting rather than desktop-first workflows.",
        specifications=json.dumps({"Deployment": "Cloud/Mobile", "Best For": "Small-Mid Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="Greenway Health (Intergy)", website="https://www.greenwayhealth.com", category="EHR/EMR",
        introduction="Greenway Health provides the Intergy EHR and practice management platform for ambulatory practices, with tools for clinical documentation, billing, and analytics.",
        ourVerdict="Greenway's Intergy is a reasonable choice for ambulatory practices wanting an established vendor with integrated billing and analytics.",
        keyTakeaways=arr("Established vendor for ambulatory practices", "Integrated practice management and analytics", "Specialty-specific content libraries"),
        pros=arr("Mature, established platform", "Good specialty content libraries", "Integrated revenue cycle tools"),
        cons=arr("Interface seen as less modern by some users", "Implementation can take significant time", "Customer support experiences vary"),
        whoIsItFor="Ambulatory practices across multiple specialties.",
        howItIsDifferent="Long-standing ambulatory focus with deep specialty-specific content libraries built up over time.",
        specifications=json.dumps({"Deployment": "Cloud/On-Premise", "Best For": "Ambulatory Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="Kareo / Tebra", website="https://www.tebra.com", category="EHR/EMR",
        introduction="Tebra (formerly Kareo, now merged with PatientPop) offers a cloud-based EHR, billing, and patient engagement platform aimed at independent practices.",
        ourVerdict="Tebra is a strong fit for independent practices wanting an all-in-one platform covering EHR, billing, and patient acquisition/marketing.",
        keyTakeaways=arr("Formed from the Kareo and PatientPop merger", "Combines EHR with marketing/patient acquisition tools", "Targets independent practices"),
        pros=arr("All-in-one EHR, billing, and marketing platform", "Easy onboarding for small practices", "Modern, user-friendly interface"),
        cons=arr("Some advanced clinical features are limited", "Pricing tiers can add up with add-ons", "Best suited to smaller practices, not large groups"),
        whoIsItFor="Independent practices wanting EHR plus marketing/patient acquisition tools.",
        howItIsDifferent="Bundles patient acquisition and marketing tools (from PatientPop) alongside the core EHR.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Independent Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="AdvancedMD", website="https://www.advancedmd.com", category="EHR/EMR",
        introduction="AdvancedMD provides cloud-based EHR, practice management, and billing software for independent medical practices of various sizes.",
        ourVerdict="AdvancedMD is well suited to independent practices that want strong billing/RCM tools paired with a configurable EHR.",
        keyTakeaways=arr("Strong combined EHR and billing/RCM offering", "Cloud-based, no on-site servers needed", "Serves practices from solo to multi-location"),
        pros=arr("Robust billing and claims management", "Good scalability for multi-location practices", "Telehealth included"),
        cons=arr("Learning curve for new users", "Customization may require vendor assistance", "Pricing can be less transparent upfront"),
        whoIsItFor="Independent practices from solo providers to multi-location groups.",
        howItIsDifferent="Emphasizes its billing/RCM engine as much as the clinical EHR itself.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Independent Practices", "Pricing Model": "Subscription + % of collections"}),
    ),
    dict(
        name="CareCloud", website="https://www.carecloud.com", category="EHR/EMR",
        introduction="CareCloud offers cloud-based EHR, practice management, and revenue cycle management solutions for medical practices.",
        ourVerdict="CareCloud is a capable option for practices wanting integrated EHR and revenue cycle services from a single vendor.",
        keyTakeaways=arr("Combines EHR with revenue cycle management services", "Cloud-based platform", "Serves practices of varying sizes"),
        pros=arr("Integrated RCM services", "Modern user interface", "Good reporting/analytics tools"),
        cons=arr("Some users report slower support response", "Implementation time can vary", "Add-on costs for premium features"),
        whoIsItFor="Practices wanting integrated EHR and revenue cycle management.",
        howItIsDifferent="Positions revenue cycle management as a core selling point alongside the EHR itself.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Independent Practices", "Pricing Model": "Subscription + % of collections"}),
    ),
    dict(
        name="ModMed (Modernizing Medicine)", website="https://www.modmed.com", category="EHR/EMR",
        introduction="ModMed provides specialty-specific EHR and practice management software, with deep workflows tailored to fields like dermatology, ophthalmology, and orthopedics.",
        ourVerdict="ModMed stands out for specialty practices that need deep, specialty-tailored workflows rather than a generic EHR.",
        keyTakeaways=arr("Recognized as a top healthcare software product in G2's 2025 Best Software Awards", "Deep specialty-specific workflows", "Strong adoption in dermatology, ophthalmology, orthopedics"),
        pros=arr("Highly tailored specialty workflows", "Strong user satisfaction in specialty markets", "Mobile-friendly clinical tools"),
        cons=arr("Less suited to general/primary care practices", "Pricing geared toward specialty practice budgets", "Switching from legacy systems takes effort"),
        whoIsItFor="Specialty practices such as dermatology, ophthalmology, gastroenterology, and orthopedics.",
        howItIsDifferent="Built from the ground up around specialty-specific clinical workflows rather than a one-size-fits-all template.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Specialty Practices", "Pricing Model": "Subscription"}),
    ),
    dict(
        name="SimplePractice", website="https://www.simplepractice.com", category="EHR/EMR",
        introduction="SimplePractice is an all-in-one EHR and practice management platform built for health and wellness providers, including therapists, counselors, and other behavioral health practitioners.",
        ourVerdict="SimplePractice is a top choice for solo and small group behavioral health and wellness practices wanting an easy, affordable EHR.",
        keyTakeaways=arr("Top-rated EHR on G2 for health and wellness providers", "Popular with behavioral health and therapy practices", "Simple setup aimed at solo/small practices"),
        pros=arr("Very user-friendly interface", "Affordable for solo/small practices", "Strong client portal and scheduling tools"),
        cons=arr("Less suited to large multi-specialty organizations", "Limited deep medical-specialty features", "Some advanced billing features require add-ons"),
        whoIsItFor="Solo and small group health and wellness practices, especially behavioral health.",
        howItIsDifferent="Focuses specifically on health and wellness/behavioral health providers rather than general medical practices.",
        specifications=json.dumps({"Deployment": "Cloud", "Best For": "Health & Wellness Providers", "Pricing Model": "Subscription"}),
    ),
]

def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Software"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in ROWS:
        row.setdefault("slug", slugify(row["name"]))
        row.setdefault("rating", "")
        row.setdefault("logo", "")
        row.setdefault("reportUrl", "")
        row.setdefault("pictures", "")
        row.setdefault("sentiments", "")
        row.setdefault("faqs", "")
        ws.append([row.get(col, "") for col in COLUMNS])

    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 28 if col not in ("introduction", "ourVerdict", "howItIsDifferent") else 50

    wb.save("D:/SoftwareDome/data/ehr-emr-software-seed.xlsx")
    print("done")

if __name__ == "__main__":
    main()
